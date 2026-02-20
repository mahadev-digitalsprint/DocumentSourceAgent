"""
M7 — Excel Writer Agent
========================
Generates a 7-sheet Excel workbook summarising all pipeline results.

Sheets:
  1. Summary         — overall stats: companies, docs, changes
  2. Financial Docs  — all financial PDFs with metadata (revenue, profit, EPS, etc.)
  3. Non-Financial   — non-financial PDFs with metadata (topics, scope, etc.)
  4. 24h Changes     — all doc & page changes in last 24 hours
  5. WebWatch        — page-level additions/deletions/content changes
  6. Metadata Raw    — full LLM JSON dump for all docs
  7. Errors          — any pipeline errors for debugging

File is saved to BASE_DOWNLOAD_PATH/reports/finwatch_<date>.xlsx
"""
import logging
import os
from datetime import datetime, timedelta
from typing import Dict, List

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl

from app.config import get_settings
from app.database import SessionLocal
from app.models import (
    Company, DocumentRegistry, MetadataRecord,
    ChangeLog, PageChange, ErrorLog
)
from app.workflow.state import PipelineState

logger = logging.getLogger(__name__)
settings = get_settings()

# ── Color palette ─────────────────────────────────────────────────────────────
COLOR_HEADER_FIN     = "1B4F72"   # dark blue  — financial sheets
COLOR_HEADER_NONFIN  = "1D6A39"   # dark green — non-financial sheets
COLOR_HEADER_CHANGE  = "7B241C"   # dark red   — changes sheet
COLOR_HEADER_NEUTRAL = "2C3E50"   # dark grey  — neutral sheets
COLOR_ALT_ROW        = "EBF5FB"   # light blue alt row


def excel_agent(state: PipelineState) -> dict:
    """LangGraph node — build and save the Excel workbook."""
    db = SessionLocal()
    try:
        report_dir = os.path.join(
            settings.base_download_path.replace("/app/downloads", "downloads"),
            "reports"
        )
        os.makedirs(report_dir, exist_ok=True)
        date_str = datetime.utcnow().strftime("%Y%m%d_%H%M")
        out_path  = os.path.join(report_dir, f"finwatch_{date_str}.xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        _sheet_summary(wb, db)
        _sheet_financial(wb, db)
        _sheet_non_financial(wb, db)
        _sheet_24h_changes(wb, db)
        _sheet_webwatch(wb, db)
        _sheet_metadata_raw(wb, db)
        _sheet_errors(wb, db)

        wb.save(out_path)
        logger.info(f"[M7-EXCEL] Saved: {out_path}")
        return {"excel_path": out_path}
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_summary(wb, db):
    ws = wb.create_sheet("📊 Summary")
    cutoff = datetime.utcnow() - timedelta(hours=24)

    companies   = db.query(Company).filter(Company.active == True).count()
    total_docs  = db.query(DocumentRegistry).count()
    fin_docs    = db.query(DocumentRegistry).filter(DocumentRegistry.doc_type.like("FINANCIAL%")).count()
    nonfin_docs = db.query(DocumentRegistry).filter(DocumentRegistry.doc_type.like("NON_FINANCIAL%")).count()
    new_24h     = db.query(ChangeLog).filter(ChangeLog.detected_at >= cutoff).count()
    errors_24h  = db.query(ErrorLog).filter(ErrorLog.created_at >= cutoff).count()

    rows = [
        ("🏢 Active Companies",         companies),
        ("📄 Total Documents",           total_docs),
        ("💰 Financial Documents",        fin_docs),
        ("📋 Non-Financial Documents",    nonfin_docs),
        ("🔔 Changes (last 24h)",         new_24h),
        ("❌ Errors (last 24h)",          errors_24h),
        ("📅 Report Generated",           datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
    ]

    _write_header(ws, ["Metric", "Value"], COLOR_HEADER_NEUTRAL)
    for i, (metric, value) in enumerate(rows, start=2):
        ws.cell(i, 1, metric)
        ws.cell(i, 2, str(value))
        if i % 2 == 0:
            for col in range(1, 3):
                ws.cell(i, col).fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25


def _sheet_financial(wb, db):
    ws = wb.create_sheet("💰 Financial Docs")
    headers = [
        "Company", "Document Type", "Headline", "Filing Date",
        "Period End", "Fiscal Year", "Fiscal Quarter", "Currency",
        "Revenue", "Net Profit", "EBITDA", "EPS",
        "Audit Status", "Preliminary", "Language", "URL", "Local Path",
    ]
    _write_header(ws, headers, COLOR_HEADER_FIN)

    docs = (
        db.query(DocumentRegistry, MetadataRecord, Company)
        .join(Company, DocumentRegistry.company_id == Company.id)
        .outerjoin(MetadataRecord, MetadataRecord.document_id == DocumentRegistry.id)
        .filter(DocumentRegistry.doc_type.like("FINANCIAL%"))
        .order_by(Company.company_name, DocumentRegistry.created_at.desc())
        .all()
    )

    for i, (doc, meta, company) in enumerate(docs, start=2):
        raw = meta.raw_llm_response if meta and meta.raw_llm_response else {}
        row = [
            company.company_name,
            (doc.doc_type or "").split("|")[-1],
            meta.headline if meta else "",
            meta.filing_date if meta else "",
            meta.period_end_date if meta else "",
            raw.get("fiscal_year", ""),
            raw.get("fiscal_quarter", ""),
            raw.get("currency", ""),
            raw.get("revenue", ""),
            raw.get("net_profit", ""),
            raw.get("ebitda", ""),
            raw.get("eps", ""),
            raw.get("audit_status", ""),
            str(raw.get("is_preliminary", "")),
            meta.language if meta else "",
            doc.document_url,
            doc.local_path or "",
        ]
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
        if i % 2 == 0:
            for col in range(1, len(headers)+1):
                ws.cell(i, col).fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

    _auto_width(ws, headers)


def _sheet_non_financial(wb, db):
    ws = wb.create_sheet("📋 Non-Financial Docs")
    headers = [
        "Company", "Document Type", "Headline", "Filing Date",
        "Regulatory Body", "Compliance Period", "Document Scope",
        "Target Audience", "Key Topics", "Key Findings",
        "Certifications", "Language", "URL",
    ]
    _write_header(ws, headers, COLOR_HEADER_NONFIN)

    docs = (
        db.query(DocumentRegistry, MetadataRecord, Company)
        .join(Company, DocumentRegistry.company_id == Company.id)
        .outerjoin(MetadataRecord, MetadataRecord.document_id == DocumentRegistry.id)
        .filter(DocumentRegistry.doc_type.like("NON_FINANCIAL%"))
        .order_by(Company.company_name, DocumentRegistry.created_at.desc())
        .all()
    )

    for i, (doc, meta, company) in enumerate(docs, start=2):
        raw = meta.raw_llm_response if meta and meta.raw_llm_response else {}
        topics = raw.get("key_topics", [])
        certs  = raw.get("certifications", [])
        row = [
            company.company_name,
            (doc.doc_type or "").split("|")[-1],
            meta.headline if meta else "",
            meta.filing_date if meta else "",
            raw.get("regulatory_body", ""),
            raw.get("compliance_period", ""),
            raw.get("document_scope", ""),
            raw.get("target_audience", ""),
            ", ".join(topics) if isinstance(topics, list) else str(topics),
            raw.get("key_findings", ""),
            ", ".join(certs) if isinstance(certs, list) else str(certs),
            meta.language if meta else "",
            doc.document_url,
        ]
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
        if i % 2 == 0:
            for col in range(1, len(headers)+1):
                ws.cell(i, col).fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

    _auto_width(ws, headers)


def _sheet_24h_changes(wb, db):
    ws = wb.create_sheet("🔔 24h Changes")
    cutoff = datetime.utcnow() - timedelta(hours=24)
    headers = [
        "Company", "Change Type", "Doc Category", "Doc Type",
        "URL", "Old Hash", "New Hash", "Detected At",
    ]
    _write_header(ws, headers, COLOR_HEADER_CHANGE)

    changes = (
        db.query(ChangeLog, DocumentRegistry, Company)
        .join(DocumentRegistry, ChangeLog.document_id == DocumentRegistry.id)
        .join(Company, DocumentRegistry.company_id == Company.id)
        .filter(ChangeLog.detected_at >= cutoff)
        .order_by(ChangeLog.detected_at.desc())
        .all()
    )

    for i, (chg, doc, company) in enumerate(changes, start=2):
        parts = (doc.doc_type or "").split("|")
        row = [
            company.company_name,
            chg.change_type,
            parts[0] if len(parts) > 1 else "UNKNOWN",
            parts[-1],
            doc.document_url,
            chg.old_hash or "",
            chg.new_hash or "",
            str(chg.detected_at)[:19],
        ]
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)

    _auto_width(ws, headers)


def _sheet_webwatch(wb, db):
    ws = wb.create_sheet("🌐 WebWatch")
    cutoff = datetime.utcnow() - timedelta(hours=24)
    headers = [
        "Company", "Page URL", "Change Type",
        "Diff Summary", "New PDFs Found", "Detected At",
    ]
    _write_header(ws, headers, COLOR_HEADER_NEUTRAL)

    pchanges = (
        db.query(PageChange, Company)
        .join(Company, PageChange.company_id == Company.id)
        .filter(PageChange.detected_at >= cutoff)
        .order_by(PageChange.detected_at.desc())
        .all()
    )

    for i, (pc, company) in enumerate(pchanges, start=2):
        new_pdfs = pc.new_pdf_urls or []
        row = [
            company.company_name,
            pc.page_url,
            pc.change_type,
            pc.diff_summary or "",
            len(new_pdfs),
            str(pc.detected_at)[:19],
        ]
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)

    _auto_width(ws, headers)


def _sheet_metadata_raw(wb, db):
    ws = wb.create_sheet("🔬 Raw Metadata")
    headers = [
        "Company", "Doc Type", "Headline", "Filing Date",
        "Period End", "Language", "Audit", "Preliminary",
        "Has Income Stmt", "Notes", "Source", "URL",
    ]
    _write_header(ws, headers, COLOR_HEADER_NEUTRAL)

    recs = (
        db.query(MetadataRecord, DocumentRegistry, Company)
        .join(DocumentRegistry, MetadataRecord.document_id == DocumentRegistry.id)
        .join(Company, DocumentRegistry.company_id == Company.id)
        .order_by(MetadataRecord.created_at.desc())
        .all()
    )

    for i, (meta, doc, company) in enumerate(recs, start=2):
        row = [
            company.company_name,
            (doc.doc_type or "").split("|")[-1],
            meta.headline or "",
            meta.filing_date or "",
            meta.period_end_date or "",
            meta.language or "",
            "Yes" if meta.audit_flag else "No",
            "Yes" if meta.preliminary_document else "No",
            "Yes" if meta.income_statement else "No",
            "Yes" if meta.note_flag else "No",
            meta.filing_data_source or "",
            doc.document_url,
        ]
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
        if i % 2 == 0:
            for col in range(1, len(headers)+1):
                ws.cell(i, col).fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

    _auto_width(ws, headers)


def _sheet_errors(wb, db):
    ws = wb.create_sheet("❌ Errors")
    headers = [
        "Company", "Step", "Error Type", "Error Message",
        "Document URL", "Created At",
    ]
    _write_header(ws, headers, "922B21")  # dark red

    errors = (
        db.query(ErrorLog, Company)
        .outerjoin(Company, ErrorLog.company_id == Company.id)
        .order_by(ErrorLog.created_at.desc())
        .limit(500)
        .all()
    )

    for i, (err, company) in enumerate(errors, start=2):
        row = [
            company.company_name if company else "N/A",
            err.step or "",
            err.error_type or "",
            (err.error_message or "")[:200],
            err.document_url or "",
            str(err.created_at)[:19],
        ]
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)

    _auto_width(ws, headers)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _write_header(ws, headers: List[str], color: str):
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor=color)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(1, j, h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 22


def _auto_width(ws, headers: List[str]):
    for i, h in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        # Set reasonable widths based on content type
        if any(kw in h.lower() for kw in ["url", "path", "message", "summary", "finding"]):
            ws.column_dimensions[col_letter].width = 40
        elif any(kw in h.lower() for kw in ["headline", "topic", "type"]):
            ws.column_dimensions[col_letter].width = 30
        else:
            ws.column_dimensions[col_letter].width = 18
